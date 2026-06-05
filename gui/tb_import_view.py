"""Trial Balance import wizard — file picker, column mapping, preview."""

from __future__ import annotations
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from config import THEME as T
from core.tb_importer import import_xlsx, import_csv, import_tally_xml
from core.tb_template_generator import generate as generate_tb_template
from gui.theme import primary_btn, secondary_btn, label


class ColumnMappingDialog(tk.Toplevel):
    """Modal dialog: shows file headers + preview rows, lets user map each column."""

    ROLES = ["— Ignore —", "Ledger Name", "Group / Mapping",
             "SubType / Sch III Sub-heading", "Debit (Dr)",
             "Credit (Cr)", "Net Balance (CY)", "Net Balance (PY)"]
    ROLE_KEYS = [None, "ledger", "group", "subtype", "debit", "credit", "net", "py_net"]

    def __init__(self, parent, headers: list[str], preview_rows: list[list],
                 auto_map: dict[str, int | None]):
        super().__init__(parent)
        self.title("Map Columns — Trial Balance Import")
        self.resizable(True, True)
        self.grab_set()
        self._headers = headers
        self._preview = preview_rows
        self._auto    = auto_map
        self._combos: list[ttk.Combobox] = []
        self.result: dict[str, int | None] | None = None
        self._build()
        self.geometry("900x520")
        self.minsize(700, 400)

    def _build(self):
        ttk.Label(self, text="Assign each column to a TB field. Ledger Name is required.",
                  style="Muted.TLabel").pack(fill="x", padx=8, pady=(8, 2))

        # Mapping row: one combobox per column
        map_frame = ttk.Frame(self)
        map_frame.pack(fill="x", padx=8, pady=4)
        for col_idx, hdr in enumerate(self._headers):
            col_frame = ttk.Frame(map_frame, relief="groove", borderwidth=1)
            col_frame.grid(row=0, column=col_idx, padx=2, pady=2, sticky="ew")
            ttk.Label(col_frame, text=hdr or f"Col {col_idx+1}",
                      font=(None, 9, "bold"), wraplength=100).pack(padx=4, pady=2)
            # Determine pre-selected role
            preselect = 0  # "— Ignore —"
            for role_key, mapped_col in self._auto.items():
                if mapped_col == col_idx:
                    try:
                        preselect = self.ROLE_KEYS.index(role_key)
                    except ValueError:
                        pass
                    break
            var = tk.StringVar(value=self.ROLES[preselect])
            cb = ttk.Combobox(col_frame, textvariable=var, values=self.ROLES,
                              state="readonly", width=14)
            cb.pack(padx=4, pady=(0, 4))
            self._combos.append(cb)
        map_frame.columnconfigure(tuple(range(len(self._headers))), weight=1)

        # Preview table
        ttk.Label(self, text="Preview (first rows):",
                  style="Muted.TLabel").pack(fill="x", padx=8, pady=(4, 0))
        prev_frame = ttk.Frame(self)
        prev_frame.pack(fill="both", expand=True, padx=8, pady=4)
        cols = [f"c{i}" for i in range(len(self._headers))]
        tv = ttk.Treeview(prev_frame, columns=cols, show="headings", height=6)
        for i, hdr in enumerate(self._headers):
            tv.heading(f"c{i}", text=hdr or f"Col {i+1}")
            tv.column(f"c{i}", width=100, minwidth=60)
        for row in self._preview:
            tv.insert("", "end", values=row[:len(self._headers)])
        vsb = ttk.Scrollbar(prev_frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")

        # Buttons
        btn = ttk.Frame(self)
        btn.pack(fill="x", padx=8, pady=6)
        from gui.theme import primary_btn, secondary_btn
        primary_btn(btn, "✔ Confirm Mapping", command=self._confirm).pack(side="right", padx=4)
        secondary_btn(btn, "Cancel", command=self.destroy).pack(side="right", padx=4)
        secondary_btn(btn, "Auto-detect", command=self._auto_detect).pack(side="left", padx=4)

    def _auto_detect(self):
        for col_idx, cb in enumerate(self._combos):
            role = 0
            for role_key, mapped_col in self._auto.items():
                if mapped_col == col_idx:
                    try:
                        role = self.ROLE_KEYS.index(role_key)
                    except ValueError:
                        pass
                    break
            cb.set(self.ROLES[role])

    def _confirm(self):
        col_map: dict[str, int | None] = {k: None for k in self.ROLE_KEYS if k}
        for col_idx, cb in enumerate(self._combos):
            chosen = cb.get()
            try:
                role_idx = self.ROLES.index(chosen)
            except ValueError:
                continue
            role_key = self.ROLE_KEYS[role_idx]
            if role_key is not None:
                col_map[role_key] = col_idx
        if col_map.get("ledger") is None:
            from tkinter import messagebox
            messagebox.showerror("Required", "Ledger Name column must be mapped.")
            return
        self.result = col_map
        self.destroy()


class TBImportView(ttk.Frame):
    def __init__(self, parent, db, on_complete: callable = None):
        super().__init__(parent)
        self._db          = db
        self._on_complete = on_complete
        self._import_result = None
        self._path: Path | None = None
        self._build()

    def _build(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=8, pady=6)
        label(top, "2.  Import Trial Balance", style="Sec.TLabel").pack(side="left")

        # Template download bar
        tmpl = ttk.Frame(self)
        tmpl.pack(fill="x", padx=8, pady=(4, 0))
        label(tmpl, "No TB file yet?").pack(side="left", padx=4)
        secondary_btn(tmpl, "📥  Download TB Template for this entity",
                      command=self._download_template).pack(side="left", padx=4)
        label(tmpl, "Fill in Excel, then import back here.",
              style="Muted.TLabel").pack(side="left", padx=6)
        secondary_btn(tmpl, "🔗  Connect Zoho Books",
                      command=self._zoho_connect).pack(side="right", padx=4)

        # File picker
        pick = ttk.Frame(self)
        pick.pack(fill="x", padx=8, pady=4)
        label(pick, "Source File:").pack(side="left", padx=4)
        self._path_var = tk.StringVar()
        ttk.Entry(pick, textvariable=self._path_var, width=50,
                  state="readonly").pack(side="left", padx=4)
        secondary_btn(pick, "Browse …", command=self._browse).pack(side="left", padx=4)
        primary_btn(pick, "Import", command=self._do_import).pack(side="left", padx=8)

        # Status
        self._status_var = tk.StringVar(value="Select a file to import Trial Balance data.")
        ttk.Label(self, textvariable=self._status_var,
                  style="Muted.TLabel", wraplength=700).pack(fill="x", padx=8, pady=2)

        # Warnings/errors text
        self._msg = tk.Text(self, height=3, bg=T["bg"], fg=T["error"],
                            font=(T["font"], 9), relief="flat", state="disabled")
        self._msg.pack(fill="x", padx=8, pady=2)

        # Preview grid
        cols = [
            ("ledger", "Ledger Name",       220, "w"),
            ("group",  "Group",             140, "w"),
            ("dr",     "Debit",             100, "e"),
            ("cr",     "Credit",            100, "e"),
            ("net",    "Closing (CY)",      110, "e"),
            ("py",     "PY Net",            110, "e"),
            ("src",    "Source",             70, "center"),
        ]
        from gui.fs_grid_view import EditableGrid
        self._grid = EditableGrid(self, columns=cols)
        self._grid.pack(fill="both", expand=True, padx=8, pady=4)

        # Bottom bar
        bot = ttk.Frame(self)
        bot.pack(fill="x", padx=8, pady=6)
        self._count_var = tk.StringVar(value="")
        ttk.Label(bot, textvariable=self._count_var,
                  style="Muted.TLabel").pack(side="left")
        primary_btn(bot, "✔ Confirm & Proceed  →", command=self._confirm).pack(side="right")
        secondary_btn(bot, "Clear / Re-import", command=self._clear).pack(side="right", padx=6)

    def _download_template(self):
        entity_type = "COMPANY"
        try:
            entity_type = self._db.get_entity_type() or "COMPANY"
        except Exception:
            pass
        entity_type = entity_type.upper()

        default_name = f"TB_Template_{entity_type}.xlsx"
        try:
            entity_name = self._db.get_entity_name() or entity_type
            fy = self._db.get_fy() or ""
            if fy:
                default_name = f"{entity_name}_TB_Template_{fy}.xlsx"
        except Exception:
            pass

        save_path = filedialog.asksaveasfilename(
            title="Save TB Template",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            generate_tb_template(entity_type, Path(save_path))
            messagebox.showinfo(
                "Template Saved",
                f"✅ TB Template saved to:\n{save_path}\n\n"
                "Open in Excel:\n"
                " • Col A — enter your ledger names\n"
                " • Col B — select the mapping from the dropdown\n"
                " • Cols C/D (or C–F for NCE) — enter amounts\n\n"
                "Then import the file back here.",
            )
        except Exception as e:
            messagebox.showerror("Template Error", f"Failed to generate template:\n{e}")

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select Trial Balance File",
            filetypes=[
                ("All supported", "*.xlsx *.xls *.csv *.txt *.xml"),
                ("Excel", "*.xlsx *.xls"),
                ("CSV / Text", "*.csv *.txt"),
                ("Tally XML", "*.xml"),
            ]
        )
        if path:
            self._path_var.set(path)
            self._path = Path(path)

    def _do_import(self):
        if not self._path or not self._path.exists():
            messagebox.showerror("Error", "Please select a valid file first.")
            return
        suffix = self._path.suffix.lower()
        try:
            if suffix in (".xlsx", ".xls"):
                # Check for FinStruct template first — skip wizard if detected
                from core.tb_importer import (
                    detect_finstruct_template, import_finstruct_template,
                    import_xlsx, import_csv, get_raw_headers_and_rows, get_auto_col_map,
                    override_columns
                )
                etype = detect_finstruct_template(self._path)
                if etype is not None:
                    result = import_finstruct_template(self._path, etype)
                else:
                    headers, preview = get_raw_headers_and_rows(self._path)
                    if not headers:
                        result = import_xlsx(self._path)
                    else:
                        result = self._run_mapping_wizard(headers, preview)
                        if result is None:
                            return  # user cancelled
            elif suffix in (".csv", ".txt"):
                from core.tb_importer import (
                    import_csv, get_raw_headers_and_rows, get_auto_col_map, override_columns
                )
                headers, preview = get_raw_headers_and_rows(self._path)
                if not headers:
                    result = import_csv(self._path)
                else:
                    result = self._run_mapping_wizard(headers, preview)
                    if result is None:
                        return
            elif suffix == ".xml":
                from core.tb_importer import import_tally_xml
                result = import_tally_xml(self._path)
            else:
                messagebox.showerror("Unsupported", f"File type '{suffix}' not supported.")
                return
        except Exception as e:
            messagebox.showerror("Import Error", str(e))
            return

        self._import_result = result
        self._render(result)

    def _run_mapping_wizard(self, headers, preview):
        from core.tb_importer import get_auto_col_map, override_columns
        from core.tb_importer import import_xlsx, import_csv
        auto_map = get_auto_col_map(headers)
        dlg = ColumnMappingDialog(self, headers, preview, auto_map)
        self.wait_window(dlg)
        if dlg.result is None:
            return None  # cancelled
        # Re-parse full file with user's column map
        suffix = self._path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(self._path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                from core.tb_importer import ImportResult
                r = ImportResult(); r.errors.append("Empty sheet"); return r
            raw_rows = all_rows[1:]
        else:
            import csv as _csv, io
            raw = self._path.read_bytes()
            text = raw.decode("utf-8-sig", errors="replace")
            import csv as _csv
            dialect = _csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            reader = _csv.reader(io.StringIO(text), dialect)
            rows = list(reader)
            raw_rows = rows[1:] if rows else []
        from core.tb_importer import ImportResult
        result = ImportResult()
        result.col_map = dlg.result
        from core.tb_importer import _parse_rows_with_map
        _parse_rows_with_map(headers, raw_rows, result, dlg.result)
        return result

    def _render(self, result):
        msgs = result.errors + result.warnings
        self._msg.configure(state="normal")
        self._msg.delete("1.0", "end")
        if msgs:
            self._msg.insert("end", "\n".join(msgs))
        self._msg.configure(state="disabled")

        grid_rows = []
        for i, row in enumerate(result.rows):
            grid_rows.append({
                "iid": str(i),
                "tag": "alt" if i % 2 else "",
                "values": [
                    row["ledger_name"], row["group_name"] or "",
                    f"{row['cy_debit']:,.2f}" if row["cy_debit"] else "—",
                    f"{row['cy_credit']:,.2f}" if row["cy_credit"] else "—",
                    f"{row['cy_net']:,.2f}"  if row["cy_net"]  else "—",
                    f"{row['py_net']:,.2f}"  if row["py_net"]  else "—",
                    row["source"],
                ],
            })
        self._grid.load_rows(grid_rows)
        n = len(result.rows)
        self._count_var.set(f"{n} ledger(s) detected  |  "
                            f"{'⚠ ' + str(len(result.warnings)) + ' warning(s)' if result.warnings else '✅ No warnings'}")
        self._status_var.set(
            f"Preview ready — {n} rows from '{self._path.name}'. "
            "Review and click Confirm.")

    def _confirm(self):
        if not self._import_result or not self._import_result.rows:
            messagebox.showerror("No Data", "No data to import. Please import a file first.")
            return
        if self._import_result.errors:
            if not messagebox.askyesno("Errors", "Import has errors. Proceed anyway?"):
                return
        self._db.clear_raw_tb()
        self._db.insert_raw_tb_batch(self._import_result.rows)

        # Apply SubType auto-mapping hints (confidence 1.0, source=SUBTYPE, auto-confirmed)
        hints = getattr(self._import_result, "subtype_hints", {}) or {}
        applied = 0
        if hints:
            raw_rows = self._db.get_raw_tb()
            ordered_ids = [r["id"] for r in raw_rows]
            for row_idx, code in hints.items():
                if row_idx >= len(ordered_ids):
                    continue
                raw_id = ordered_ids[row_idx]
                src_row = self._import_result.rows[row_idx]
                self._db.upsert_wtb(
                    raw_id, code, 1.0, "SUBTYPE",
                    src_row["cy_net"], src_row["py_net"], confirmed=1,
                )
                applied += 1

        self._db.log("TB_IMPORTED",
                     f"{len(self._import_result.rows)} rows from {self._path.name}"
                     + (f"; auto-mapped {applied} via SubType" if applied else ""))
        messagebox.showinfo("Imported",
                            f"✅ {len(self._import_result.rows)} ledgers imported."
                            + (f"\n✨ {applied} auto-mapped from SubType column."
                               if applied else ""))
        if self._on_complete:
            self._on_complete()

    def _clear(self):
        self._import_result = None
        self._path = None
        self._path_var.set("")
        self._grid.load_rows([])
        self._count_var.set("")
        self._status_var.set("Select a file to import Trial Balance data.")
        self._msg.configure(state="normal")
        self._msg.delete("1.0", "end")
        self._msg.configure(state="disabled")

    def _zoho_connect(self):
        from gui.zoho_connect_dialog import ZohoConnectDialog
        ZohoConnectDialog(self, self._db)
