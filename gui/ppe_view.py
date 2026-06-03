"""PPE Register GUI — asset grid + depreciation computation."""

from __future__ import annotations
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from openpyxl import load_workbook
from ..config import THEME as T, PPE_CATEGORIES
from ..core.ppe_engine import recalc_asset, summarize_ppe
from ..core.ppe_template_generator import generate_ppe_template
from ..gui.theme import primary_btn, secondary_btn, label
from .fs_grid_view import EditableGrid


class PPEView(ttk.Frame):
    def __init__(self, parent, db, on_dep_posted: callable = None):
        super().__init__(parent)
        self._db          = db
        self._on_dep_posted = on_dep_posted
        self._build()
        self._load()

    def _build(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=8, pady=6)
        label(top, "5.  PPE / Fixed Asset Register", style="Sec.TLabel").pack(side="left")
        primary_btn(top, "+ Add Asset", command=self._add_asset).pack(side="left", padx=8)
        secondary_btn(top, "📥  Import from Excel", command=self._import_ppe).pack(side="left", padx=4)
        secondary_btn(top, "📥  Download Template", command=self._download_template).pack(side="left", padx=4)
        secondary_btn(top, "Post Depreciation to WTB", command=self._post_dep).pack(side="left", padx=4)
        secondary_btn(top, "Delete Selected", command=self._delete_asset).pack(side="left", padx=4)

        cols = [
            ("asset",    "Asset Description",   200, "w"),
            ("cat",      "Category",             140, "w"),
            ("method",   "Method",                60, "center"),
            ("life",     "Life (Yrs)",             65, "center"),
            ("gross_op", "Gross Blk Op",          100, "e"),
            ("adds",     "Additions",              100, "e"),
            ("disp",     "Disposals",              90, "e"),
            ("gross_cl", "Gross Blk Cl",           100, "e"),
            ("dep_op",   "Acc Dep Op",             100, "e"),
            ("dep_ch",   "Dep Charge",             100, "e"),
            ("dep_cl",   "Acc Dep Cl",             100, "e"),
            ("nbv_cy",   "Net Block CY",           100, "e"),
            ("nbv_py",   "Net Block PY",           100, "e"),
            ("it_dep",   "IT Dep",                  80, "e"),
        ]
        self._grid = EditableGrid(
            self, columns=cols,
            on_cell_change=self._on_change,
            editable_cols={"asset","cat","method","life","gross_op",
                           "adds","disp","dep_op","dep_ch","dep_cl","nbv_py","it_dep"}
        )
        self._grid.pack(fill="both", expand=True, padx=8, pady=4)

        # Totals bar
        self._tot_var = tk.StringVar(value="")
        ttk.Label(self, textvariable=self._tot_var,
                  style="Muted.TLabel").pack(fill="x", padx=8, pady=2)

    def _load(self):
        rows_db = self._db.get_ppe()
        self._assets = []
        grid_rows = []
        for i, a in enumerate(rows_db):
            d = dict(a)
            self._assets.append(d)
            grid_rows.append(self._make_row(d, i))
        self._grid.load_rows(grid_rows)
        self._refresh_totals()

    def _make_row(self, a: dict, i: int) -> dict:
        def f(k): return f"{float(a.get(k) or 0):,.2f}"
        return {
            "iid":    str(a.get("id", i)),
            "tag":    "alt" if i % 2 else "",
            "values": [
                a.get("asset_name",""), a.get("category",""), a.get("method","SLM"),
                a.get("useful_life_yrs","10"),
                f("gross_op"), f("additions"), f("disposals"), f("gross_cl"),
                f("dep_op"), f("dep_charge"), f("dep_cl"), f("nbv_cy"),
                f("nbv_py"), f("it_dep"),
            ]
        }

    def _refresh_totals(self):
        tot = summarize_ppe(self._assets)
        self._tot_var.set(
            f"Total Net Block CY: ₹{tot['nbv_cy']:,.2f}  |  "
            f"Total Net Block PY: ₹{tot['nbv_py']:,.2f}  |  "
            f"Total Dep Charge: ₹{tot['dep_charge']:,.2f}  |  "
            f"Total IT Dep: ₹{tot['it_dep']:,.2f}"
        )

    def _add_asset(self):
        new = {"asset_name": "New Asset", "category": PPE_CATEGORIES[0],
               "method": "SLM", "useful_life_yrs": 10}
        self._db.upsert_ppe(new)
        self._load()

    def _delete_asset(self):
        iid = self._grid.get_selected_iid()
        if not iid:
            messagebox.showinfo("Select Row", "Please select an asset to delete.")
            return
        if messagebox.askyesno("Delete", "Delete selected asset?"):
            self._db.delete_ppe(int(iid))
            self._load()

    def _recalc_all(self):
        rows = self._grid.get_all_rows()
        for i, row in enumerate(rows):
            if i >= len(self._assets):
                break
            a = self._assets[i]
            a["asset_name"]     = row[0]
            a["category"]       = row[1]
            a["method"]         = row[2]
            a["useful_life_yrs"]= int(row[3] or 10)
            for j, key in enumerate(["gross_op","additions","disposals","dep_op"], start=4):
                try:
                    a[key] = float(str(row[j]).replace(",",""))
                except (ValueError, IndexError):
                    a[key] = 0.0
            r = recalc_asset(a)
            self._assets[i] = r
            self._db.upsert_ppe(r)
        self._load()

    # Entity-type-aware depreciation account mapping (Dr expense / Cr asset-dep)
    DEP_CODES = {
        "COMPANY": ("PL025", "AS002", "PL026", "AS005"),  # tang_dep, acc_dep, intang_amort, acc_amort
        "SEC8":    ("PL025", "AS002", "PL026", "AS005"),
        "LLP":     ("LL025", "LL010", "LL025", "LL011"),
        "PROP":    ("NP008", "NC012", "NP008", "NC013"),
        "PART":    ("NP008", "NC012", "NP008", "NC013"),
        "AOP":     ("AE004", "AO009", "AE004", "AO009"),
        "TRUST":   ("TE004", "TR007", "TE004", "TR007"),
    }

    def _is_intangible(self, asset: dict) -> bool:
        cat = (asset.get("category", "") or "").lower()
        return "intangible" in cat or "software" in cat or "goodwill" in cat

    def _post_dep(self):
        tot = summarize_ppe(self._assets)
        dep = tot["dep_charge"]
        if dep == 0:
            messagebox.showinfo("No Depreciation",
                                "No depreciation calculated yet.\n"
                                "Add asset rows with gross_op / useful_life filled in first.")
            return

        et = (self._db.get_meta("entity_type") or "COMPANY").upper()
        dep_codes = self.DEP_CODES.get(et, self.DEP_CODES["COMPANY"])
        tang_dr, tang_cr, intang_dr, intang_cr = dep_codes

        # Split tangible vs intangible
        tang_dep = sum(float(a.get("dep_charge", 0) or 0)
                       for a in self._assets if not self._is_intangible(a))
        intang_dep = sum(float(a.get("dep_charge", 0) or 0)
                         for a in self._assets if self._is_intangible(a))

        # Confirm if re-posting
        existing = [a for a in self._db.get_adjustments()
                    if a["adj_id"] and a["adj_id"].startswith("DEP-")]
        if existing:
            if not messagebox.askyesno(
                "Re-post Depreciation",
                f"Found {len(existing)} prior depreciation adjustment(s).\n"
                f"Replace with fresh entries totalling ₹{dep:,.2f}?"):
                return
            self._db.delete_dep_adjustments()

        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        entries = []
        if tang_dep > 0.01:
            self._db.add_adjustment(
                f"DEP-TANG-DR-{ts}", "Depreciation on Tangible Assets",
                tang_dr, tang_dep, 0, "Depreciation per PPE Register (Tangible)")
            self._db.add_adjustment(
                f"DEP-TANG-CR-{ts}", "Accumulated Depreciation",
                tang_cr, 0, tang_dep, "Depreciation per PPE Register (Tangible)")
            entries.append(f"Tangible: ₹{tang_dep:,.2f}")
        if intang_dep > 0.01:
            self._db.add_adjustment(
                f"DEP-INT-DR-{ts}", "Amortisation of Intangible Assets",
                intang_dr, intang_dep, 0, "Amortisation per PPE Register")
            self._db.add_adjustment(
                f"DEP-INT-CR-{ts}", "Accumulated Amortisation",
                intang_cr, 0, intang_dep, "Amortisation per PPE Register")
            entries.append(f"Intangible: ₹{intang_dep:,.2f}")

        self._db.log("DEP_POSTED", f"Total ₹{dep:,.2f}; entity={et}")
        messagebox.showinfo(
            "Posted",
            f"✅ Depreciation posted (Total ₹{dep:,.2f}):\n  " + "\n  ".join(entries)
            + f"\n\nEntity: {et}\n"
            + f"Dr {tang_dr} / Cr {tang_cr} (tangible)"
            + (f"\nDr {intang_dr} / Cr {intang_cr} (intangible)" if intang_dep > 0 else "")
        )
        if self._on_dep_posted:
            self._on_dep_posted(dep)

    def _on_change(self, iid: str, col_id: str, new_val: str):
        pass  # No auto-recalc; user imports final figures

    def _download_template(self):
        save_path = filedialog.asksaveasfilename(
            title="Save PPE Data Entry Template",
            defaultextension=".xlsx",
            initialfile="PPE_Template.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            generate_ppe_template(Path(save_path))
            messagebox.showinfo(
                "Template Saved",
                f"✅ PPE template saved to:\n{save_path}\n\n"
                "Instructions:\n"
                " • Fill Col A: Asset ID (e.g., LA001)\n"
                " • Fill Col B: Asset Name\n"
                " • Fill Col C: Category from list\n"
                " • Fill Cols D-H: Gross/Accumulated Dep values for CY/PY and CWIP\n\n"
                "Then use 'Import from Excel' to load the data."
            )
        except Exception as e:
            messagebox.showerror("Template Error", f"Failed to generate template:\n{e}")

    def _import_ppe(self):
        path = filedialog.askopenfilename(
            title="Select PPE Data File",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")]
        )
        if not path:
            return
        try:
            wb = load_workbook(Path(path), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=4, values_only=True))
            wb.close()

            # Parse rows: AssetID, Name, Category, Gross CY, Acc Dep CY, Gross PY, Acc Dep PY, CWIP
            imported = 0
            for row in rows:
                if not row or not row[0]:
                    continue  # Skip empty rows
                try:
                    asset = {
                        "asset_id": str(row[0] or "").strip(),
                        "asset_name": str(row[1] or "").strip(),
                        "category": str(row[2] or "Plant & Machinery").strip(),
                        "method": "SLM",  # Default
                        "useful_life_yrs": 5,  # Default
                        "gross_op": float(row[5] or 0) if len(row) > 5 else 0,  # Gross PY
                        "gross_cl": float(row[3] or 0) if len(row) > 3 else 0,  # Gross CY
                        "additions": 0,  # Not in template
                        "disposals": 0,   # Not in template
                        "dep_op": float(row[6] or 0) if len(row) > 6 else 0,   # Acc Dep PY
                        "dep_cl": float(row[4] or 0) if len(row) > 4 else 0,   # Acc Dep CY
                        "dep_charge": 0,  # Will be computed if needed
                        "nbv_py": 0,      # Will be computed
                        "nbv_cy": 0,      # Will be computed
                        "it_dep": 0,      # Default
                    }
                    if asset["asset_name"]:
                        self._db.upsert_ppe(asset)
                        imported += 1
                except (ValueError, IndexError, TypeError) as e:
                    continue

            if imported > 0:
                self._load()
                messagebox.showinfo("Imported",
                    f"✅ {imported} asset(s) imported successfully from PPE template.")
            else:
                messagebox.showwarning("No Data", "No valid asset data found in file.")
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import PPE data:\n{e}")
