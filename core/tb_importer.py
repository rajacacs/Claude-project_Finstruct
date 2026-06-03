"""Trial Balance importer — XLSX, CSV, Tally XML, and FinStruct templates."""

from __future__ import annotations
import csv
import io
import logging
import re
from pathlib import Path

log = logging.getLogger(__name__)

COMMON_LEDGER_HEADERS = {"ledger", "account", "particulars", "name", "description",
                          "account name", "accountname"}
COMMON_GROUP_HEADERS  = {"group", "category", "parent", "type", "nature", "accounttype"}
COMMON_SUBTYPE_HEADERS = {"subtype", "sub type", "sub-type", "sch iii sub heading",
                          "sub-heading", "sub heading", "mapping"}
COMMON_DR_HEADERS     = {"debit", "dr", "debit amount", "debit balance", "closing debit"}
COMMON_CR_HEADERS     = {"credit", "cr", "credit amount", "credit balance", "closing credit"}
COMMON_NET_HEADERS    = {"net", "closing", "balance", "amount", "closing balance",
                          "net amount", "cy net"}
COMMON_PYNET_HEADERS  = {"py", "previous", "prev year", "last year", "prior year",
                          "py amount", "py net", "prev year figure", "previous year"}

BALANCING_LINE_PATTERN = re.compile(
    r"^(current\s+year\s+)?(loss|profit|surplus|deficit)(\s+(transfer|for\s+the\s+year))?$|"
    r"^p\s*&?\s*l\s+transfer|"
    r"^profit\s*&?\s*loss\s+transfer",
    re.IGNORECASE,
)


def _norm(s: str) -> str:
    return str(s or "").strip().lower().replace("_", " ")


def _detect_col(headers: list[str], candidates: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if _norm(h) in candidates:
            return i
    return None


def _strip_dr_cr_text(s: str) -> str:
    """Remove Dr./Cr. suffix tokens used by Tally in closing balance text."""
    for tag in ("Dr.", "Cr.", "DR.", "CR.", "Dr", "dr", "CR", "Cr", "cr"):
        s = s.replace(tag, "")
    return s


def _to_float(v) -> float:
    try:
        s = str(v or "").strip()
        # Strip currency prefixes
        for prefix in ("₹", "Rs.", "Rs", "INR"):
            s = s.replace(prefix, "")
        s = s.strip()
        # Handle Dr/Cr suffix — extract sign first, then strip suffix
        suffix_sign = 1
        sl = s.lower()
        for sfx in ("cr", "credit"):
            if sl.endswith(sfx):
                suffix_sign = -1
                s = s[: -len(sfx)].strip()
                break
        else:
            for sfx in ("dr", "debit"):
                if sl.endswith(sfx):
                    s = s[: -len(sfx)].strip()
                    break
        # Handle bracket notation for negatives: (1234) → -1234
        s = s.replace(",", "").replace("(", "-").replace(")", "")
        val = float(s) if s else 0.0
        return val * suffix_sign
    except (ValueError, TypeError):
        return 0.0


def _is_balancing_line(ledger_name: str) -> bool:
    """Detect the 'Current Year Loss / P&L Transfer' balancing row in synthetic TBs."""
    name = (ledger_name or "").strip()
    if not name:
        return False
    if BALANCING_LINE_PATTERN.search(name):
        return True
    nl = name.lower()
    return ("balancing figure" in nl) or ("p&l transfer" in nl) or ("p & l transfer" in nl)


def _normalise(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"[\(\)\–\-/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _build_subtype_index() -> dict:
    """Build a multi-level lookup index from MASTER.

    Returns dict with keys:
      'sub'  : sub_heading.lower() → code
      'head' : heading.lower()     → list of (code, sub_heading)
      'all'  : list of MappingEntry (for keyword substring fallback)
    """
    from .master_db import MASTER
    sub_idx: dict[str, str] = {}
    head_idx: dict[str, list] = {}
    entries = []
    for m in MASTER:
        entries.append(m)
        sub_l = _normalise(m.sub_heading)
        if sub_l and sub_l not in sub_idx:
            sub_idx[sub_l] = m.code
        head_l = _normalise(m.heading)
        head_idx.setdefault(head_l, []).append((m.code, m.sub_heading))
    return {"sub": sub_idx, "head": head_idx, "all": entries}


def _lookup_subtype_code(subtype: str, idx: dict, account_name: str = "") -> str | None:
    """Multi-tier match.

    1. Exact match on sub_heading
    2. Match heading + use account_name to disambiguate sub_heading within heading
    3. Substring/keyword match across MASTER sub_headings
    """
    if not subtype:
        return None
    s = _normalise(subtype)
    sub_idx = idx["sub"]
    head_idx = idx["head"]

    # Tier 1: exact sub_heading match
    if s in sub_idx:
        return sub_idx[s]

    # Tier 2: matches a heading → use account_name to pick the right sub_heading
    if s in head_idx:
        candidates = head_idx[s]
        an = _normalise(account_name)
        if an:
            # Find candidate whose sub_heading appears in account_name
            best: tuple[int, str] | None = None
            for code, sub in candidates:
                sub_n = _normalise(sub)
                # Score by longest substring intersection
                if sub_n and sub_n in an:
                    score = len(sub_n)
                    if not best or score > best[0]:
                        best = (score, code)
            if best:
                return best[1]
            # Try reverse: account_name keywords appearing in sub_heading
            for code, sub in candidates:
                sub_n = _normalise(sub)
                for word in an.split():
                    if len(word) >= 4 and word in sub_n:
                        return code
        # Fall back to first candidate in this heading
        return candidates[0][0]

    # Tier 3: substring match on account_name vs sub_headings
    an = _normalise(account_name)
    if an:
        best: tuple[int, str] | None = None
        for sub_n, code in sub_idx.items():
            if not sub_n:
                continue
            if sub_n in an:
                if not best or len(sub_n) > best[0]:
                    best = (len(sub_n), code)
        if best:
            return best[1]

    # Tier 4: substring keyword on subtype itself
    for sub_n, code in sub_idx.items():
        if (s in sub_n or sub_n in s) and abs(len(s) - len(sub_n)) < 12:
            return code

    return None


class ImportResult:
    def __init__(self):
        self.rows: list[dict] = []
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.col_map: dict[str, int | None] = {}
        # SubType → mapping_code hints (auto-confirmed at confidence 1.0)
        self.subtype_hints: dict[int, str] = {}


def detect_finstruct_template(path: Path) -> str | None:
    """Check cell A1 of 'TrialBalance' sheet for a FinStruct sentinel; return entity_type or None."""
    try:
        from openpyxl import load_workbook
        from .tb_template_generator import SENTINELS
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["TrialBalance"] if "TrialBalance" in wb.sheetnames else wb.active
        val = str(ws.cell(1, 1).value or "").strip()
        wb.close()
        for etype, sentinel in SENTINELS.items():
            if val == sentinel:
                return etype
    except Exception as e:
        log.debug("Template detection failed: %s", e)
    return None


def import_finstruct_template(path: Path, entity_type: str) -> ImportResult:
    """Import a FinStruct standardised TB template (row 1 = sentinel, row 2 = headers, data from row 3).

    Net-balance types (COMPANY, SEC8, LLP): col A=ledger, col B=lookup_name, col C=cy_net, col D=py_net
    Dr/Cr types (PROP, PART, AOP, TRUST):   col A=ledger, col B=group,       col C=cy_dr, col D=cy_cr,
                                                                               col E=py_dr, col F=py_cr
    """
    result = ImportResult()
    _NET_BALANCE = {"COMPANY", "SEC8", "LLP"}
    etype = entity_type.upper()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["TrialBalance"] if "TrialBalance" in wb.sheetnames else wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        result.errors.append(f"Template read error: {e}")
        return result

    # Rows: 0=sentinel, 1=headers, 2+ = data
    if len(all_rows) < 3:
        result.errors.append("Template has no data rows.")
        return result

    is_net = etype in _NET_BALANCE
    source_tag = f"TEMPLATE_{etype}"

    for raw_row in all_rows[2:]:
        row = list(raw_row)
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        ledger = str(row[0] if row else "").strip()
        if not ledger:
            continue
        mapping = str(row[1] if len(row) > 1 else "").strip()

        if is_net:
            cy_net = _to_float(row[2] if len(row) > 2 else 0)
            py_net = _to_float(row[3] if len(row) > 3 else 0)
            cy_dr  = cy_net if cy_net >= 0 else 0.0
            cy_cr  = abs(cy_net) if cy_net < 0 else 0.0
        else:
            cy_dr  = _to_float(row[2] if len(row) > 2 else 0)
            cy_cr  = _to_float(row[3] if len(row) > 3 else 0)
            py_dr  = _to_float(row[4] if len(row) > 4 else 0)
            py_cr  = _to_float(row[5] if len(row) > 5 else 0)
            cy_net = cy_dr - cy_cr
            py_net = py_dr - py_cr

        if not mapping:
            result.warnings.append(f"Row '{ledger}': no mapping selected — skipped.")
            continue

        result.rows.append({
            "ledger_name": ledger,
            "group_name":  mapping,
            "cy_debit":    cy_dr,
            "cy_credit":   cy_cr,
            "cy_net":      cy_net,
            "py_net":      py_net,
            "source":      source_tag,
        })

    if not result.rows:
        result.errors.append("No data rows found in template (check rows start from row 3).")
    return result


def import_xlsx(path: Path) -> ImportResult:
    # Check if this is a FinStruct template first
    etype = detect_finstruct_template(path)
    if etype is not None:
        return import_finstruct_template(path, etype)

    result = ImportResult()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            result.errors.append("Empty sheet")
            return result
        headers = [str(c or "") for c in all_rows[0]]
        _parse_rows(headers, all_rows[1:], result, source="XLSX")
        wb.close()
    except Exception as e:
        result.errors.append(f"XLSX read error: {e}")
    return result


def import_csv(path: Path) -> ImportResult:
    result = ImportResult()
    try:
        raw = path.read_bytes()
        text = raw.decode("utf-8-sig", errors="replace")
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
        reader = csv.reader(io.StringIO(text), dialect)
        rows = list(reader)
        if not rows:
            result.errors.append("Empty CSV")
            return result
        headers = rows[0]
        _parse_rows(headers, rows[1:], result, source="CSV")
    except Exception as e:
        result.errors.append(f"CSV read error: {e}")
    return result


def get_raw_headers_and_rows(path: Path) -> tuple[list[str], list[list]]:
    """Return (headers, first_8_rows) for column mapping wizard."""
    suffix = path.suffix.lower()
    try:
        if suffix in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return [], []
            headers = [str(c or "") for c in all_rows[0]]
            data = [[str(v or "") for v in row] for row in all_rows[1:9]]
            return headers, data
        elif suffix in (".csv", ".txt"):
            import csv as _csv, io as _io
            raw = path.read_bytes()
            text = raw.decode("utf-8-sig", errors="replace")
            dialect = _csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            reader = _csv.reader(_io.StringIO(text), dialect)
            rows = list(reader)
            if not rows:
                return [], []
            return rows[0], rows[1:9]
    except Exception:
        pass
    return [], []


def get_auto_col_map(headers: list[str]) -> dict[str, int | None]:
    return {
        "ledger": _detect_col(headers, COMMON_LEDGER_HEADERS),
        "group":  _detect_col(headers, COMMON_GROUP_HEADERS),
        "debit":  _detect_col(headers, COMMON_DR_HEADERS),
        "credit": _detect_col(headers, COMMON_CR_HEADERS),
        "net":    _detect_col(headers, COMMON_NET_HEADERS),
        "py_net": _detect_col(headers, COMMON_PYNET_HEADERS),
    }


def import_tally_xml(path: Path) -> ImportResult:
    result = ImportResult()
    try:
        import xml.etree.ElementTree as ET
        tree = ET.parse(path)
        root = tree.getroot()
        for ledger in root.iter("LEDGER"):
            name = ledger.get("NAME") or (ledger.findtext("NAME") or "").strip()
            parent = ledger.findtext("PARENT") or ""
            cl_bal_txt = ledger.findtext("CLOSINGBALANCE") or "0"
            norm_bal = cl_bal_txt.lower().replace(".", "").strip()
            is_dr = "dr" in norm_bal
            net = _to_float(_strip_dr_cr_text(cl_bal_txt))
            if not is_dr:
                net = -net
            op_bal_txt = ledger.findtext("OPENINGBALANCE") or "0"
            op_norm    = op_bal_txt.lower().replace(".", "").strip()
            op_is_dr   = "dr" in op_norm
            py_net     = _to_float(_strip_dr_cr_text(op_bal_txt))
            if not op_is_dr:
                py_net = -py_net
            if name:
                result.rows.append({
                    "ledger_name": name,
                    "group_name":  parent,
                    "cy_debit":    net if net >= 0 else 0,
                    "cy_credit":   abs(net) if net < 0 else 0,
                    "cy_net":      net,
                    "py_net":      py_net,
                    "source":      "XML",
                })
    except Exception as e:
        result.errors.append(f"XML parse error: {e}")
    return result


def _parse_rows(headers: list[str], data_rows, result: ImportResult, source: str = "XLSX"):
    lcol  = _detect_col(headers, COMMON_LEDGER_HEADERS)
    gcol  = _detect_col(headers, COMMON_GROUP_HEADERS)
    stcol = _detect_col(headers, COMMON_SUBTYPE_HEADERS)
    dcol  = _detect_col(headers, COMMON_DR_HEADERS)
    ccol  = _detect_col(headers, COMMON_CR_HEADERS)
    ncol  = _detect_col(headers, COMMON_NET_HEADERS)
    pcol  = _detect_col(headers, COMMON_PYNET_HEADERS)

    result.col_map = {
        "ledger": lcol, "group": gcol, "subtype": stcol, "debit": dcol,
        "credit": ccol, "net": ncol, "py_net": pcol,
    }

    if lcol is None:
        result.errors.append(
            "Could not detect Ledger column. Please map columns manually."
        )
        lcol = 0

    # Build subtype index always — even without explicit SubType column,
    # we can attempt account-name → sub_heading matching as a fallback.
    subtype_idx = _build_subtype_index()
    skipped_balancing = 0

    for i, row in enumerate(data_rows, start=2):
        row = list(row)
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        name = str(row[lcol] if lcol < len(row) else "").strip()
        if not name:
            continue
        if _is_balancing_line(name):
            skipped_balancing += 1
            continue
        group  = str(row[gcol] if gcol is not None and gcol < len(row) else "").strip()
        subtyp = str(row[stcol] if stcol is not None and stcol < len(row) else "").strip()
        dr     = _to_float(row[dcol]) if dcol is not None and dcol < len(row) else 0.0
        cr     = _to_float(row[ccol]) if ccol is not None and ccol < len(row) else 0.0
        net    = _to_float(row[ncol]) if ncol is not None and ncol < len(row) else (dr - cr)
        py     = _to_float(row[pcol]) if pcol is not None and pcol < len(row) else 0.0
        result.rows.append({
            "ledger_name": name,
            "group_name":  group or subtyp,
            "cy_debit":    dr,
            "cy_credit":   cr,
            "cy_net":      net,
            "py_net":      py,
            "source":      source,
        })
        if subtype_idx and (subtyp or name):
            code = _lookup_subtype_code(subtyp, subtype_idx, account_name=name)
            if code:
                result.subtype_hints[len(result.rows) - 1] = code

    if skipped_balancing:
        result.warnings.append(
            f"Skipped {skipped_balancing} 'Current Year Loss / P&L Transfer' balancing row(s)."
        )
    if result.subtype_hints:
        result.warnings.append(
            f"Auto-mapped {len(result.subtype_hints)} ledger(s) using SubType column."
        )

    seen: dict[str, int] = {}
    for r in result.rows:
        n = r["ledger_name"].lower()
        seen[n] = seen.get(n, 0) + 1
    for name, cnt in seen.items():
        if cnt > 1:
            result.warnings.append(f"Duplicate ledger: '{name}' appears {cnt} times")


def override_columns(
    import_result: ImportResult,
    raw_rows: list,
    headers: list[str],
    col_map: dict[str, int | None],
) -> ImportResult:
    """Re-parse with user-supplied column assignments."""
    result = ImportResult()
    result.col_map = col_map
    _parse_rows_with_map(headers, raw_rows, result, col_map)
    return result


def _parse_rows_with_map(headers, data_rows, result, col_map):
    lcol  = col_map.get("ledger")
    gcol  = col_map.get("group")
    stcol = col_map.get("subtype")
    dcol  = col_map.get("debit")
    ccol  = col_map.get("credit")
    ncol  = col_map.get("net")
    pcol  = col_map.get("py_net")
    if lcol is None:
        result.errors.append("Ledger column not mapped")
        return
    # Build subtype index always — even without explicit SubType column,
    # we can attempt account-name → sub_heading matching as a fallback.
    subtype_idx = _build_subtype_index()
    skipped_balancing = 0
    for row in data_rows:
        row = list(row)
        name = str(row[lcol] if lcol < len(row) else "").strip()
        if not name:
            continue
        if _is_balancing_line(name):
            skipped_balancing += 1
            continue
        group  = str(row[gcol] if gcol is not None and gcol < len(row) else "").strip()
        subtyp = str(row[stcol] if stcol is not None and stcol < len(row) else "").strip()
        dr  = _to_float(row[dcol]) if dcol is not None and dcol < len(row) else 0.0
        cr  = _to_float(row[ccol]) if ccol is not None and ccol < len(row) else 0.0
        net = _to_float(row[ncol]) if ncol is not None and ncol < len(row) else (dr - cr)
        py  = _to_float(row[pcol]) if pcol is not None and pcol < len(row) else 0.0
        result.rows.append({
            "ledger_name": name, "group_name": group or subtyp,
            "cy_debit": dr, "cy_credit": cr, "cy_net": net, "py_net": py, "source": "MANUAL",
        })
        if subtype_idx and (subtyp or name):
            code = _lookup_subtype_code(subtyp, subtype_idx, account_name=name)
            if code:
                result.subtype_hints[len(result.rows) - 1] = code
    if skipped_balancing:
        result.warnings.append(
            f"Skipped {skipped_balancing} 'Current Year Loss / P&L Transfer' balancing row(s)."
        )
    if result.subtype_hints:
        result.warnings.append(
            f"Auto-mapped {len(result.subtype_hints)} ledger(s) using SubType column."
        )
