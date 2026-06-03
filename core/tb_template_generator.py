"""Entity-specific TB template generator.

Generates XLSX templates aligned with ICAI-prescribed formats for each entity type,
derived from the GAS reference implementations (NCE automators + Sch III tool).
"""

from __future__ import annotations
from pathlib import Path

from .master_db import get_master, MappingEntry

SENTINEL_PREFIX = "__FINSTRUCT_TB_TEMPLATE_"
SENTINEL_SUFFIX = "__"

# Canonical entity type → sentinel string
SENTINELS: dict[str, str] = {
    "COMPANY": f"{SENTINEL_PREFIX}COMPANY{SENTINEL_SUFFIX}",
    "SEC8":    f"{SENTINEL_PREFIX}SEC8{SENTINEL_SUFFIX}",
    "LLP":     f"{SENTINEL_PREFIX}LLP{SENTINEL_SUFFIX}",
    "PROP":    f"{SENTINEL_PREFIX}PROP{SENTINEL_SUFFIX}",
    "PART":    f"{SENTINEL_PREFIX}PART{SENTINEL_SUFFIX}",
    "AOP":     f"{SENTINEL_PREFIX}AOP{SENTINEL_SUFFIX}",
    "TRUST":   f"{SENTINEL_PREFIX}TRUST{SENTINEL_SUFFIX}",
}

# Entity types using 4-column (net balance) format: Ledger | Mapping lookup_name | CY Net | PY Net
_NET_BALANCE_TYPES = {"COMPANY", "SEC8", "LLP"}

# Entity types using 6-column (Dr/Cr) format: Ledger | Group | CY Dr | CY Cr | PY Dr | PY Cr
_DR_CR_TYPES = {"PROP", "PART", "AOP", "TRUST"}

# Header colours matching GAS tool conventions
_HDR_BG   = "1F3864"
_HDR_FG   = "FFFFFF"
_SEC_BG   = "D6E4F0"
_SEC_FG   = "1F3864"
_ALT_BG   = "F2F9FF"
_TOT_BG   = "1F3864"
_AMT_FMT  = '#,##0.00'


def detect_template(path: Path) -> str | None:
    """Read cell A1 of 'TrialBalance' sheet; return entity_type string or None."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["TrialBalance"] if "TrialBalance" in wb.sheetnames else wb.active
        val = str(ws.cell(1, 1).value or "").strip()
        wb.close()
        for etype, sentinel in SENTINELS.items():
            if val == sentinel:
                return etype
    except Exception:
        pass
    return None


def generate(entity_type: str, output_path: Path) -> None:
    """Generate an entity-specific XLSX TB template and save to output_path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise ImportError("openpyxl is required for TB template generation") from exc

    etype = entity_type.upper()
    sentinel = SENTINELS.get(etype)
    if sentinel is None:
        raise ValueError(f"Unknown entity type: {entity_type!r}")

    wb = Workbook()
    wb.remove(wb.active)

    # Determine format type
    is_net_balance = etype in _NET_BALANCE_TYPES

    # Build master entries for this entity
    tag_map = {
        "COMPANY": ["COMPANY"], "SEC8": ["SEC8", "COMPANY"],
        "LLP":     ["LLP"],
        "PROP":    ["PROP"],    "PART": ["PART"],
        "AOP":     ["AOP"],     "TRUST": ["TRUST", "NPO"],
    }
    tags = tag_map.get(etype, [etype])
    entries: list[MappingEntry] = get_master(tags)

    # ── Sheet 1: Instructions ────────────────────────────────────────────────
    ws_instr = wb.create_sheet("Instructions")
    _build_instructions(ws_instr, etype, is_net_balance, Font, PatternFill, Alignment)

    # ── Sheet 2: MappingReference ────────────────────────────────────────────
    ws_ref = wb.create_sheet("MappingReference")
    lookup_names = _build_mapping_reference(ws_ref, entries, is_net_balance,
                                             Font, PatternFill, Alignment, Border, Side)

    # ── Sheet 3: TrialBalance ────────────────────────────────────────────────
    ws_tb = wb.create_sheet("TrialBalance")
    _build_trial_balance(ws_tb, sentinel, etype, is_net_balance, lookup_names,
                          Font, PatternFill, Alignment, Border, Side, DataValidation)

    wb.save(output_path)


# ── Internal builders ────────────────────────────────────────────────────────

def _build_instructions(ws, etype, is_net_balance, Font, PatternFill, Alignment):
    hdr_fill = PatternFill("solid", fgColor=_HDR_BG)
    hdr_font = Font(bold=True, color=_HDR_FG, size=11)
    body_font = Font(size=10)

    ws.column_dimensions["A"].width = 100

    rows = [
        ("FINSTRUCT — TRIAL BALANCE IMPORT TEMPLATE", True),
        (f"Entity Type: {etype}", False),
        ("", False),
        ("HOW TO USE THIS TEMPLATE", True),
        ("", False),
        ("Step 1 — Open the 'TrialBalance' sheet.", False),
        ("Step 2 — In Column A, enter your ledger names exactly as they appear in your books of account.", False),
        (f"Step 3 — In Column B, select the {'mapping lookup path' if is_net_balance else 'mapping group'} "
         f"that best matches each ledger from the dropdown list.", False),
        ("         (The full list of valid mappings is in the 'MappingReference' sheet for reference.)", False),
    ]
    if is_net_balance:
        rows += [
            ("Step 4 — In Column C, enter the NET closing balance for the Current Year (positive = debit, negative = credit).", False),
            ("Step 5 — In Column D, enter the NET closing balance for the Previous Year.", False),
        ]
    else:
        rows += [
            ("Step 4 — Enter Current Year Debit balance in Column C.", False),
            ("Step 5 — Enter Current Year Credit balance in Column D.", False),
            ("Step 6 — Enter Previous Year Debit balance in Column E.", False),
            ("Step 7 — Enter Previous Year Credit balance in Column F.", False),
            ("         Note: Enter only the Dr OR Cr side for each ledger. Leave the other column blank/zero.", False),
        ]
    rows += [
        ("", False),
        ("Step — Save the file and import it back into FinStruct via the 'Import TB' option.", False),
        ("       FinStruct will auto-detect this template and skip the column-mapping wizard.", False),
        ("", False),
        ("IMPORTANT: Do not modify row 1 (sentinel) or row 2 (headers) of the TrialBalance sheet.", False),
        ("           Do not rename or delete this sheet.", False),
        ("           Amounts should be in Indian Rupees (₹). Do not include commas; the app handles formatting.", False),
    ]

    for i, (text, bold) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, color=_HDR_FG if i == 1 else "201F1E",
                             size=12 if i == 1 else 10)
            if i == 1:
                cell.fill = hdr_fill
        else:
            cell.font = body_font
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 18


def _build_mapping_reference(ws, entries: list[MappingEntry], is_net_balance,
                               Font, PatternFill, Alignment, Border, Side) -> list[str]:
    hdr_fill = PatternFill("solid", fgColor=_HDR_BG)
    hdr_font = Font(bold=True, color=_HDR_FG, size=10)
    alt_fill = PatternFill("solid", fgColor=_ALT_BG)

    if is_net_balance:
        headers = ["Lookup Name (paste into TrialBalance Col B)", "Code", "BS / PL", "Note #", "Sign Convention"]
        col_widths = [70, 10, 8, 8, 25]
    else:
        headers = ["Mapping Group (select in TrialBalance Col B)", "Code", "BS / IE / RP", "Note #", "Sign Convention"]
        col_widths = [50, 10, 10, 8, 25]

    for col, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    lookup_names = []
    seen: set[str] = set()

    for row_idx, entry in enumerate(entries, start=2):
        key = entry.lookup_name if is_net_balance else entry.group
        if key in seen:
            continue
        seen.add(key)
        lookup_names.append(key)

        sign_label = "Debit = positive" if entry.sign == "DR_POSITIVE" else "Credit = positive"
        row_data = [key, entry.code, entry.fs_tag,
                    str(entry.note_number) if entry.note_number else "—", sign_label]
        fill = alt_fill if (row_idx % 2 == 0) else None
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            cell.border = border
        ws.row_dimensions[row_idx].height = 16

    return lookup_names


def _build_trial_balance(ws, sentinel, etype, is_net_balance, lookup_names: list[str],
                          Font, PatternFill, Alignment, Border, Side, DataValidation):
    hdr_fill = PatternFill("solid", fgColor=_HDR_BG)
    hdr_font = Font(bold=True, color=_HDR_FG, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin, right=thin)

    DATA_ROWS = 500

    # Row 1: sentinel
    ws.cell(row=1, column=1, value=sentinel).font = Font(size=8, color="999999")

    # Row 2: column headers
    if is_net_balance:
        headers = [
            "Ledger Name (as per Books)",
            "Mapping — Schedule III / LLP Lookup Name",
            f"Closing Balance — CY (₹)\n(positive = Dr, negative = Cr)",
            "Previous Year Figure (₹)",
        ]
        col_widths = [40, 65, 24, 24]
        amt_cols = [3, 4]
    else:
        headers = [
            "Ledger Name (as per Books)",
            "Mapping Group",
            "Current Year — Dr (₹)",
            "Current Year — Cr (₹)",
            "Previous Year — Dr (₹)",
            "Previous Year — Cr (₹)",
        ]
        col_widths = [40, 40, 18, 18, 18, 18]
        amt_cols = [3, 4, 5, 6]

    for col, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # Freeze first two rows and first column
    ws.freeze_panes = "B3"

    # Data rows
    alt_fill = PatternFill("solid", fgColor=_ALT_BG)
    for row in range(3, 3 + DATA_ROWS):
        fill = alt_fill if (row % 2 == 0) else None
        n_cols = len(headers)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(size=10)
            cell.border = border
            if fill:
                cell.fill = fill
            if col in amt_cols:
                cell.number_format = _AMT_FMT
                cell.alignment = Alignment(horizontal="right")
        ws.row_dimensions[row].height = 16

    # Data validation dropdown for Col B (mapping)
    if lookup_names:
        # openpyxl data validation with a formula referencing MappingReference sheet
        # Use a named range approach for large lists
        max_ref_row = len(lookup_names) + 1
        dv = DataValidation(
            type="list",
            formula1=f"MappingReference!$A$2:$A${max_ref_row}",
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Select a valid mapping from the dropdown list."
        dv.errorTitle = "Invalid Mapping"
        dv.prompt = "Select the FS line / group this ledger belongs to."
        dv.promptTitle = "Mapping"
        ws.add_data_validation(dv)
        dv.add(f"B3:B{3 + DATA_ROWS - 1}")

    # Totals row at the bottom
    total_row = 3 + DATA_ROWS
    ws.cell(row=total_row, column=1, value="— CHECK TOTALS —").font = Font(bold=True, size=10)
    for col in amt_cols:
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({col_letter}3:{col_letter}{total_row - 1})")
        ws.cell(row=total_row, column=col).number_format = _AMT_FMT
        ws.cell(row=total_row, column=col).font = Font(bold=True, size=10)
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="FFF9C4")


def get_column_letter(col: int) -> str:
    """Convert 1-based column index to Excel letter (A, B, ... Z, AA, ...)."""
    result = ""
    while col:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result
