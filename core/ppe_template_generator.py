"""Generate standardized PPE (Property, Plant & Equipment) data entry templates as XLSX."""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_ppe_template(output_path: Path) -> None:
    """Generate a PPE data entry template XLSX file.

    Template structure:
    - Row 1: Header with instructions
    - Row 2: Column headers
    - Rows 3+: Data entry rows (50 blank rows for asset entries)

    Columns: AssetID | Name | Category | Gross CY | Acc. Dep. CY | Gross PY | Acc. Dep. PY | CWIP
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PPE Data"

    # Header styling
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF0078D4")
    hdr_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="FFEDEBE9"),
        right=Side(style="thin", color="FFEDEBE9"),
        top=Side(style="thin", color="FFEDEBE9"),
        bottom=Side(style="thin", color="FFEDEBE9")
    )

    # Title row
    ws.append(["Fixed Assets (PPE) Data Entry Template"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(name="Calibri", bold=True, size=12)
    ws.row_dimensions[1].height = 22

    # Instructions row
    ws.append(["Enter asset details below. Category auto-fills depreciation assumptions."])
    ws.merge_cells("A2:H2")
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="FF605E5C")

    # Column headers (row 3)
    headers = [
        "Asset ID",
        "Asset Name",
        "Category",
        "Gross Value (CY)",
        "Accumulated Dep. (CY)",
        "Gross Value (PY)",
        "Accumulated Dep. (PY)",
        "CWIP (CY)"
    ]
    ws.append(headers)

    # Format header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_alignment
        cell.border = border
    ws.row_dimensions[3].height = 24

    # Set column widths
    col_widths = [12, 25, 18, 18, 18, 18, 18, 14]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Add category dropdown options and data validation to Category column
    category_options = [
        "Land",
        "Buildings",
        "Plant & Machinery",
        "Vehicles",
        "Office Equipment",
        "IT/Software",
        "Furniture & Fixtures",
        "Others"
    ]

    # Add 50 blank data rows with formatting
    for row_num in range(4, 54):
        ws.append([""] * 8)
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.font = Font(name="Calibri", size=10)
            # Right-align numeric columns
            if col_num in (4, 5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Add category dropdown data validation (for visual reference, not enforced)
    # Category options listed on a separate sheet
    cat_sheet = wb.create_sheet("Categories")
    for idx, cat in enumerate(category_options, 1):
        cat_sheet[f"A{idx}"] = cat

    # Instructions sheet
    inst_sheet = wb.create_sheet("Instructions")
    inst_sheet.column_dimensions["A"].width = 80
    instructions = [
        ("PPE Template — Data Entry Instructions", Font(bold=True, size=12)),
        ("", None),
        ("1. Asset ID", Font(bold=True)),
        ("   Unique identifier for each asset (e.g., LA001, BL001, PM001)", None),
        ("", None),
        ("2. Asset Name", Font(bold=True)),
        ("   Description of the asset (e.g., 'Land in Bangalore', 'Office Building')", None),
        ("", None),
        ("3. Category", Font(bold=True)),
        ("   Select from list: Land, Buildings, Plant & Machinery, Vehicles, etc.", None),
        ("   (See 'Categories' sheet for full list)", None),
        ("", None),
        ("4. Gross Value (CY / PY)", Font(bold=True)),
        ("   Historical cost of the asset as on Current Year / Previous Year", None),
        ("", None),
        ("5. Accumulated Depreciation (CY / PY)", Font(bold=True)),
        ("   Total depreciation accumulated as on Current Year / Previous Year", None),
        ("", None),
        ("6. CWIP (CY)", Font(bold=True)),
        ("   Capital Work-in-Progress — ongoing projects not yet capitalized", None),
        ("", None),
        ("NOTE: Net Block = Gross Value - Accumulated Depreciation + CWIP", None),
        ("      Do NOT include CWIP in the depreciation calculation.", None),
    ]

    for row_num, (text, font_style) in enumerate(instructions, 1):
        cell = inst_sheet[f"A{row_num}"]
        cell.value = text
        if font_style:
            cell.font = font_style
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    inst_sheet.row_dimensions[1].height = 22

    wb.save(output_path)
